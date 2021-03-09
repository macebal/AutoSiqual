import openpyxl
from datetime import datetime, timedelta

def buscar_fecha(fecha, hoja):
    for row in hoja.iter_rows(min_col=0, max_col=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                if cell.value == fecha:
                    return cell.row
    else:
        return -1


def buscar_fecha_mas_cercana(fecha, hoja, tipoBusqueda):
    # tipobusqueda es 0 para buscar fechas anteriores o 1 para buscar posteriores
    fechaBusqueda = fecha
    resultado = buscar_fecha(fecha, hoja)

    if resultado == -1:
        if tipoBusqueda == 0:
            fechaBusqueda -= timedelta(days=1)
        else:
            fechaBusqueda += timedelta(days=1)

        return buscar_fecha_mas_cercana(fechaBusqueda, hoja, tipoBusqueda)
    else:
        return resultado


def buscar_columna_parametro(parametro, hoja):
    for row in hoja.iter_rows(min_row=4,max_row=4):
        for cell in row:
            if cell.value == parametro:
                return cell.column - 1 #es -1 porque la columna 4 se convierte en 3 cuando la utilizas en el array de filas
    else:
        return -1
    

def importar_datos(fechaInicio, cemento):

    wb = openpyxl.load_workbook('\\\\10.15.24.32\\FC_UN_PROC y CALID\\01-Calidad\Base de Datos Productos\\Base de datos análisis de productos.xlsx', read_only=True)

    nombre_hojas_BDAP = {
                    "CPC40": "CPC40(D)",
                    "CPC30": "CPC30 (D) ",
                    "CPN40": "CPN40(D)",
                    "Plasticor": "PLASTICOR (D)"} #El nombre de las hojas de los cementos en la base de datos

    nombre_columnas_BDAP = {"Fecha": "Fecha",
                         "CAOT": "CaO",
                         "SIO2": "SiO2",
                         "AL2O3": "Al2O3",
                         "FE2O3": "Fe2O3",
                         "MGO": "MgO",
                         "SO3": "SO3",
                         "K2O": "K2O",
                         "NA2O": "Na2O",
                         "CAOL": "Cal libre",
                         "CO2": "CO2",
                         "PF": "P.P.C.",
                         "RICARB": "R.I.",
                         "G45U": "R45 m  (%)",
                         "G75U": "R75 m  (%)",
                         "SBA": "Blaine (m2/kg)",
                         "MVOL": "Densidad Real (g/cm3)",
                         "IP": "Ppio. Fragüe",
                         "FP": "Fin. Fragüe",
                         "EXAU": "Exp. (%)",
                         "AGP": "Agua PN (%)",
                         "BARI": "P Litro (g/l)",
                         "R2": "2 días", 
                         "R7": "7 días"}

    
    ws = wb[nombre_hojas_BDAP[cemento]]

    #creo un diccionario de la forma "RICARB" -> 12 donde 12 es la columna donde esta ese parametro. es None si no existe
    columnas = {}
    for key in nombre_columnas_BDAP:
        columna = buscar_columna_parametro(nombre_columnas_BDAP[key],ws)
        columnas[key] = columna if columna != -1 else None
    
    fechaFin = datetime(datetime.now().year, datetime.now().month, datetime.now().day) - timedelta(days=10)  # La fecha de fin tentativa es 10 dias antes de la fecha actual

    filaFin = buscar_fecha_mas_cercana(fechaFin, ws, 0)  # Buscar a partir de la fechaFin tentativa la fecha mas cercana que exista
    filaInicio = buscar_fecha_mas_cercana(fechaInicio, ws, 0) + 1 #la fila que le sigue a la de la ultima fecha cargada

    if filaFin > filaInicio:
        listaDatos = [] #lista vacia

        #solo itero el rango necesario para acelerar la ejecución
        indice = 0
        for fila in ws.iter_rows(min_row = filaInicio, max_row = filaFin):

            dictDatosHora = {}

            for k in nombre_columnas_BDAP:
                if columnas.get(k) != None:
                    if str(fila[columnas.get(k)].value).strip() != '*':
                        #Si existe la columna y no hay un asterisco, lo agrego al diccionario. Sino va None.
                        dictDatosHora[k] = fila[columnas.get(k)].value
                    else:
                        dictDatosHora[k] = None
                else:
                    dictDatosHora[k] = None
                       
            listaDatos.append(dictDatosHora)
            
        wb.close()
        return listaDatos
    else:
        raise ValueError('No hay valores para cargar')
